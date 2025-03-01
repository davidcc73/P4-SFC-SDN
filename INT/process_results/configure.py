from cmath import sqrt
import pprint
import constants
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font



def get_line_column_to_copy_from(sheet_to_copy_from_name, variable_number):
    line =None
    col = None

    workbook = load_workbook(constants.final_file_path)
    sheet_to_copy_from = workbook[sheet_to_copy_from_name]

    variable_name = constants.headers_lines[variable_number]

    pass_1_occurance = True          #there are 2 Lines on collumn A that have the same name
    if variable_number == 14:
        pass_1_occurance = False 

    # sheet_to_copy_from, get the line of the cell that contains the variable_name on collumn A and the collumn after it
    for row in sheet_to_copy_from.iter_rows(min_row=1, max_row=sheet_to_copy_from.max_row, min_col=1, max_col=1):
        
        if pass_1_occurance == False and row[0].value == "AVG 1º Packet Delay (nanoseconds)":
            pass_1_occurance = True
            continue
        
        if variable_number <= 9:
            if row[0].value == variable_name:
                # Get the next collumn letter of the cell that contains the variable_name
                line = row[0].row
                col = get_column_letter(row[0].column + 1)
                break
        elif variable_number == 10 or variable_number == 12:
            if row[0].value == "Mean":
                line = row[0].row
                if variable_number == 10:
                    col = get_column_letter(row[0].column + 1)
                else:
                    col = get_column_letter(row[0].column + 2)
                break
        elif variable_number == 11 or variable_number == 13:
            if row[0].value == "Standard Deviation":
                line = row[0].row
                if variable_number == 11:
                    col = get_column_letter(row[0].column + 1)
                else:
                    col = get_column_letter(row[0].column + 2)
                break
        elif variable_number == 14:
            if row[0].value == "AVG 1º Packet Delay (nanoseconds)":
                line = row[0].row
                col = get_column_letter(row[0].column + 3)
                break
        elif variable_number == 15:
            if row[0].value == "AVG Flow Delay (nanoseconds)":
                line = row[0].row
                col = get_column_letter(row[0].column + 3)
                break

    return line, col

def get_byte_sum(start, end, dscp, dscp_condition):
    # Initialize the result dictionary to store total byte counts per switch ID
    sum = {}
    switch_ids = list(range(1, constants.num_switches + 1))

    # Loop through each unique switch ID
    for switch_id in switch_ids:
        # Formulate the query to get the sum of bytes for the given switch ID and time range
        query = f"""
            SELECT SUM("size") AS total_count
            FROM flow_stats 
            WHERE time >= '{start}' AND time <= '{end}' 
            AND path =~ /(^|-)({switch_id})(-|$|\b)/
            {dscp_condition}
        """

        result = constants.apply_query(query)  # Assume this returns a dictionary with 'total_count'
        
        # Add the result to the sum dictionary under the switch_id key and dscp key
        if dscp not in sum:
            sum[dscp] = {}
        if switch_id not in sum[dscp]:
            sum[dscp][switch_id] = {}
        
        if not result.raw["series"]:            #if there is no data, set to 0
            sum[dscp][switch_id]["Byte Sums"] = 0
        else:
            sum[dscp][switch_id]["Byte Sums"] = result.raw["series"][0]["values"][0][1]

    return sum

def calculate_percentages(start, end, switch_data, dscp, dscp_condition):
    # Get the total count of packets
    query = f"""
        SELECT COUNT("latency") AS total_count
        FROM flow_stats
        WHERE time >= '{start}' 
        AND time <= '{end}'
        {dscp_condition}
    """
    result = constants.apply_query(query)
    total_count = result.raw["series"][0]["values"][0][1]  # Extract total_count from the result

    # Get the count of packets that went to each switch
    query = f"""
        SELECT COUNT("latency") AS switch_count
        FROM switch_stats
        WHERE time >= '{start}'
        AND time <= '{end}'
        {dscp_condition}
        GROUP BY switch_id
    """
    result = constants.apply_query(query)
    
    # Calculate the percentage of packets that went to each switch
    # initialize to all switches as 0, so unused switches are taken into account too
    for switch_id in range(1, constants.num_switches + 1):
        switch_data[dscp][switch_id]["Percentage Pkt"] = 0

    for row in result.raw["series"]:
        #tuple pair: id, count
        switch_id = int(row["tags"]["switch_id"])
        switch_count = int(row["values"][0][1])
        switch_data[dscp][switch_id]["Percentage Pkt"] = round((switch_count / total_count) * 100, 2)

    #pprint(switch_data)
    return switch_data

def get_mean_standard_deviation(switch_data, dscp):
    sum_percentage = 0
    sum_byte = 0
    count = 0

    # Get the mean of the (Byte Sums) and (Percentage Pkt) for all switches in switch_data
    for switch_id in switch_data[dscp]:
        sum_percentage += switch_data[dscp][switch_id]["Percentage Pkt"]
        sum_byte       += switch_data[dscp][switch_id]["Byte Sums"]
        count += 1

    percentage_mean = sum_percentage / count
    byte_mean = sum_byte / count
    

    # Get the Standard Deviation of the (Byte Sums) and (Percentage Pkt) for all switches in switch_data
    sum_squared_diff_percentage = 0
    sum_squared_diff_byte = 0
    
    for switch_id in switch_data[dscp]:
        current_percentage = switch_data[dscp][switch_id]["Percentage Pkt"]
        current_byte       = switch_data[dscp][switch_id]["Byte Sums"]
        sum_squared_diff_percentage += (current_percentage - percentage_mean) ** 2
        sum_squared_diff_byte       += (current_byte - byte_mean) ** 2
    
    percentage_std_dev = sqrt(sum_squared_diff_percentage / count).real
    byte_std_dev = sqrt(sum_squared_diff_byte / count).real
    
    switch_data[dscp]["Percentage Mean"]               = round(percentage_mean, 2)
    switch_data[dscp]["Byte Mean"]                     = round(byte_mean, 2)
    switch_data[dscp]["Percentage Standard Deviation"] = round(percentage_std_dev, 2)
    switch_data[dscp]["Byte Standard Deviation"]       = round(byte_std_dev, 2)

    return switch_data

def write_INT_results_switchID(sheet, switch_data, dscp):
    # Write the results in the sheet
    last_line = sheet.max_row + 1

    # Set new Switch ID headers
    if dscp == -1:
        title = "Switch ID For All Flows"
    else:
        title = f"Switch ID For Flows with DSCP = {dscp}"

    sheet[f'A{last_line + 1}'] = title
    sheet[f'B{last_line + 1}'] = "% of packets to each switch"
    sheet[f'C{last_line + 1}'] = "Total Sum of Processed Bytes"

    sheet[f'A{last_line + 1}'].font = Font(bold=True)
    sheet[f'B{last_line + 1}'].font = Font(bold=True)
    sheet[f'C{last_line + 1}'].font = Font(bold=True)


    # Write percentages and total bytes processed, cycle through keys that are numbers
    for i, switch_id in enumerate(switch_data[dscp].keys()):
        if isinstance(switch_id, int):                #skip sets that are non-switch_id
            sheet[f'A{last_line + 2 + i}'] = switch_id
            
            #percentage of total packets that went to each switch
            sheet[f'B{last_line + 2 + i}'] = switch_data[dscp][switch_id]["Percentage Pkt"]
            
            #Sum of processed bytes
            sheet[f'C{last_line + 2 + i}'] = switch_data[dscp][switch_id]["Byte Sums"]

    # Write the mean and standard deviation of the percentages and bytes
    sheet[f'A{last_line + constants.num_switches + 1 + 1}'] = "Mean"
    sheet[f'A{last_line + constants.num_switches + 1 + 2}'] = "Standard Deviation"
    sheet[f'A{last_line + constants.num_switches + 1 + 1}'].font = Font(bold=True)
    sheet[f'A{last_line + constants.num_switches + 1 + 2}'].font = Font(bold=True)

    sheet[f'B{last_line + constants.num_switches + 1 + 1}'] = switch_data[dscp]["Percentage Mean"]
    sheet[f'B{last_line + constants.num_switches + 1 + 2}'] = switch_data[dscp]["Percentage Standard Deviation"]
    sheet[f'C{last_line + constants.num_switches + 1 + 1}'] = switch_data[dscp]["Byte Mean"]
    sheet[f'C{last_line + constants.num_switches + 1 + 2}'] = switch_data[dscp]["Byte Standard Deviation"]

def write_INT_results(sheet, AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency, dscp):
    if dscp == -1:
        aux_dscp = "All Flows"
    else:
        aux_dscp = dscp

    # Write the results in the sheet
    last_line = sheet.max_row + 1

    # Set new Calculation headers
    sheet[f'A{last_line + 0}'] = "AVG Flows Latency (nanoseconds)"
    sheet[f'A{last_line + 1}'] = "STD Flows Latency (nanoseconds)"
    sheet[f'A{last_line + 2}'] = "AVG Hop Latency (nanoseconds)"
    sheet[f'A{last_line + 3}'] = "STD Hop Latency (nanoseconds)"

    sheet[f'A{last_line + 0}'].font = Font(bold=True)
    sheet[f'A{last_line + 1}'].font = Font(bold=True)
    sheet[f'A{last_line + 2}'].font = Font(bold=True)
    sheet[f'A{last_line + 3}'].font = Font(bold=True)

    sheet[f'B{last_line + 0}'] = AVG_flows_latency
    sheet[f'B{last_line + 1}'] = STD_flows_latency
    sheet[f'B{last_line + 2}'] = AVG_hop_latency
    sheet[f'B{last_line + 3}'] = STD_hop_latency

    sheet[f'C{last_line + 0}'] = aux_dscp
    sheet[f'C{last_line + 1}'] = aux_dscp
    sheet[f'C{last_line + 2}'] = aux_dscp
    sheet[f'C{last_line + 3}'] = aux_dscp


def set_pkt_loss():

    # Configure each sheet
    workbook = load_workbook(constants.final_file_path)

    # Set formula for each sheet
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]

        #Set new headers
        sheet['M1'] = "Packet Loss"
        sheet['N1'] = "Packet Loss (%)"

        sheet['M1'].font = Font(bold=True)
        sheet['N1'].font = Font(bold=True)

        no_formula_section = False

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            if row[0].value == "Calculations":
                break

            if row[0].value and row[0].value.startswith("Iteration -"):
                no_formula_section = False
            if no_formula_section:
                continue

            #if cell from collumn A does not contain an IPv4 address, skip
            if row[0].value is None or "." not in row[0].value:
                skip = True
                continue
            if skip:            #not in the right line of the pair
                skip = False
                continue
            
            # Set the formula, pkt loss, -1 is sender, 0 is receiver
            sheet[f'M{row[0].row}'] = f'=H{row[0].row-1}-H{row[0].row}'     
            sheet[f'N{row[0].row}'] = f'=ROUND((M{row[0].row}/H{row[0].row-1})*100, 2)'

            skip = True

    # Save the workbook
    workbook.save(constants.final_file_path)

def set_fist_pkt_delay():

    # Configure each sheet
    workbook = load_workbook(constants.final_file_path)
    
    # Set formula for each sheet
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]

        #Set new headers as bold text
        sheet['O1'] = "1º Packet Delay (nanoseconds)"
        sheet['O1'].font = Font(bold=True)

        no_formula_section = False
        
        # Set collumn L to contain a formula to be the subtraction of values of collum F of the current pair of lines
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            if row[0].value == "Calculations":
                break

            if row[0].value and row[0].value.startswith("Iteration -"):
                no_formula_section = False
            if no_formula_section:
                continue

            #if cell from collumn A does not contain an IPv4 address, skip
            if row[0].value is None or "." not in row[0].value:
                skip = True
                continue
            if skip:            #not in the right line of the pair
                skip = False
                continue
            #print(row)
            
            # Set the formula, pkt loss, -1 is sender, 0 is receiver
            # The values are 2 Timestamp (seconds-Unix Epoch)
            # subtraction give seconds, we convert to nanoseconds
            sheet[f'O{row[0].row}'] = f'=ROUND((I{row[0].row}-I{row[0].row-1})*10^9, 2)'     

            skip = True

    # Save the workbook
    workbook.save(constants.final_file_path)

def set_caculation_formulas(dscp):
    if dscp == -1:
        title = "Calculations For All Flows"
        condition = "\">0\""
        aux_dscp = "All Flows"
    else:
        title = f"Calculations For Flows with DSCP = {dscp}"
        condition = dscp
        aux_dscp = dscp
    # Configure each sheet
    workbook = load_workbook(constants.final_file_path)

    # Set formula for each sheet
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]

        #Pass the last line with data, and leave 2 empty lines
        last_line = sheet.max_row + 4

        #Set new headers
        sheet[f'A{last_line}'] = title
        sheet[f'A{last_line + 1}'] = "AVG Out of Order Packets (Nº)"
        sheet[f'A{last_line + 2}'] = "AVG Packet Loss (Nº)"
        sheet[f'A{last_line + 3}'] = "AVG Packet Loss (%)"
        sheet[f'A{last_line + 4}'] = "AVG 1º Packet Delay (nanoseconds)"
        sheet[f'A{last_line + 5}'] = "AVG Flow Jitter (nanoseconds)"
        sheet[f'A{last_line + 6}'] = "STD Flow Jitter (nanoseconds)"
        sheet[f'B{last_line}'] = "Values"
        sheet[f'C{last_line}'] = "DSCP"

        sheet[f'A{last_line}'].font = Font(bold=True)
        sheet[f'A{last_line + 1}'].font = Font(bold=True)
        sheet[f'A{last_line + 2}'].font = Font(bold=True)
        sheet[f'A{last_line + 3}'].font = Font(bold=True)
        sheet[f'A{last_line + 4}'].font = Font(bold=True)
        sheet[f'A{last_line + 5}'].font = Font(bold=True)
        sheet[f'A{last_line + 6}'].font = Font(bold=True)
        sheet[f'B{last_line}'].font = Font(bold=True)
        sheet[f'C{last_line}'].font = Font(bold=True)

        # on the next line for each column, set the average of the column, ignore empty cells
        sheet[f'B{last_line + 1}'] = f'=ROUND(AVERAGEIF(E1:E{constants.last_line_data}, {condition}, J1:J{constants.last_line_data}), 2)'
        sheet[f'B{last_line + 2}'] = f'=ROUND(AVERAGEIF(E1:E{constants.last_line_data}, {condition}, M1:M{constants.last_line_data}), 2)'
        sheet[f'B{last_line + 3}'] = f'=ROUND(AVERAGEIF(E1:E{constants.last_line_data}, {condition}, N1:N{constants.last_line_data}), 2)'
        sheet[f'B{last_line + 4}'] = f'=ROUND(AVERAGEIF(E1:E{constants.last_line_data}, {condition}, O1:O{constants.last_line_data}), 2)'
        sheet[f'B{last_line + 5}'] = f'=ROUND(AVERAGEIF(E1:E{constants.last_line_data}, {condition}, L1:L{constants.last_line_data}), 2)'
        sheet[f'B{last_line + 6}'] = constants.aux_calculated_results[dscp]["std_jitter"]       #array formuals are not working, so we calculated and set the value here

        sheet[f'C{last_line + 1}'] = aux_dscp
        sheet[f'C{last_line + 2}'] = aux_dscp
        sheet[f'C{last_line + 3}'] = aux_dscp
        sheet[f'C{last_line + 4}'] = aux_dscp
        sheet[f'C{last_line + 5}'] = aux_dscp
        sheet[f'C{last_line + 6}'] = aux_dscp

    # Save the workbook
    workbook.save(constants.final_file_path)

def get_avg_stdev_flow_hop_latency(start, end, dscp_condition):
    ############################################ Get the results from the DB
    # We need AVG Latency of ALL flows combined (NOT distinguishing between flows)
    # Query to get the 95th percentile latency value, to exclude outliers
    percentile_query = f"""
        SELECT PERCENTILE("latency", 95) AS p_latency
        FROM flow_stats
        WHERE time >= '{start}'
        AND time <= '{end}'
        {dscp_condition}
    """

    percentile_result = constants.apply_query(percentile_query)
    p_latency = list(percentile_result.get_points())[0]['p_latency']                #nanoseconds

    query = f"""
                SELECT MEAN("latency"), STDDEV("latency")
                FROM  flow_stats
                WHERE time >= '{start}'
                AND time <= '{end}'
                AND "latency" <= {p_latency}
                {dscp_condition}
            """
    result = constants.apply_query(query)
    AVG_flows_latency = round(result.raw["series"][0]["values"][0][1], 2)           #nanoseconds
    STD_flows_latency = round(result.raw["series"][0]["values"][0][2], 2)

    ###########################################
    # We need AVG Latency for processing of ALL packets (NOT distinguishing between switches/flows) 
    # Query to get the 95th percentile latency value, to exclude outliers
    percentile_query = f"""
        SELECT PERCENTILE("latency", 95) AS p_latency
        FROM switch_stats
        WHERE time >= '{start}'
        AND time <= '{end}'
        {dscp_condition}
    """
    
    query = f"""
                SELECT MEAN("latency"), STDDEV("latency")
                FROM  switch_stats
                WHERE time >= '{start}'
                AND time <= '{end}'
                AND "latency" <= {p_latency}
                {dscp_condition}
            """
    result = constants.apply_query(query)
    AVG_hop_latency = round(result.raw["series"][0]["values"][0][1], 2)             #nanoseconds
    STD_hop_latency = round(result.raw["series"][0]["values"][0][2], 2)         

    return AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency

def set_INT_results(dscp):
    # For each sheet and respectice file, see the time interval given, get the values from the DB, and set the values in the sheet
    
    if dscp == -1:
        dscp_condition = ""
    else:
        dscp_condition = f"AND dscp = \'{dscp}\'"

    # Configure each sheet
    workbook = load_workbook(constants.final_file_path)

    # Get nº each sheet
    for i, sheet in enumerate(workbook.sheetnames):

        #can i can not exceed the number of args.f (last one is comparasions)
        if i >= len(constants.args.f):
            break

        print(f"Processing sheet {sheet},\t index {i},\t for dscp {dscp}")
        sheet = workbook[sheet]

        # Get the start and end times
        start = constants.args.start[i]
        end = constants.args.end[i]

        #get the flow and hop latency, for the given dscp, that includes all flows and switches
        AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency = get_avg_stdev_flow_hop_latency(start, end, dscp_condition)

        # % of packets that went to each individual switch (switch_id)
        switch_data = get_byte_sum(start, end, dscp, dscp_condition)
        switch_data = calculate_percentages(start, end, switch_data, dscp, dscp_condition)
        switch_data = get_mean_standard_deviation(switch_data, dscp)

        write_INT_results(sheet, AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency, dscp)
        write_INT_results_switchID(sheet, switch_data, dscp)

        # Save the workbook
        workbook.save(constants.final_file_path)

def set_caculation_section(dscp):
    set_caculation_formulas(dscp)
    set_INT_results(dscp)               #technically, also contains another section, but its easier to call it here

def get_flow_delays(start, end):
    # Get the average delay of emergency and non-emergency flows
    query = f"""
        SELECT MEAN("latency") 
        FROM  flow_stats WHERE time >= '{start}' 
        AND time <= '{end}' 
        AND dscp >= 40
    """
    result = constants.apply_query(query)
    if not result.raw["series"]:
        avg_emergency_flows_delay = "none"
    else:
        avg_emergency_flows_delay = round(result.raw["series"][0]["values"][0][1], 2)         #nanoseconds

    query = f"""
        SELECT MEAN("latency")
        FROM  flow_stats
        WHERE time >= '{start}' AND time <= '{end}'
        AND dscp < 40
    """

    result = constants.apply_query(query)
    if not result.raw["series"]:
        avg_non_emergency_flows_delay = "none"
    else:
        avg_non_emergency_flows_delay = round(result.raw["series"][0]["values"][0][1], 2)         #nanoseconds

    return avg_emergency_flows_delay, avg_non_emergency_flows_delay 

def set_compare_non_Emergency_to_Emergency_variation():
    # Configure each sheet
    workbook = load_workbook(constants.final_file_path)

    for i, sheet in enumerate(workbook.sheetnames):
        #can i can not exceed the number of args.f (last one is comparasions)
        if i >= len(constants.args.f):
            break

        sheet = workbook[sheet]
        sheet.append([""])
        sheet.append([""])
        sheet.append([""])
        sheet.append([""])

        # Set new headers
        max_line = sheet.max_row
        sheet[f'A{max_line + 2}'] = "Flows Types"
        sheet[f'B{max_line + 2}'] = "Non-Emergency Flows"
        sheet[f'C{max_line + 2}'] = "Emergency Flows"
        sheet[f'D{max_line + 2}'] = "Variation (%)"
        
        sheet[f'A{max_line + 3}'] = "AVG 1º Packet Delay (nanoseconds)"
        sheet[f'A{max_line + 4}'] = "AVG Flow Delay (nanoseconds)"

        sheet[f'A{max_line + 2}'].font = Font(bold=True)
        sheet[f'B{max_line + 2}'].font = Font(bold=True)
        sheet[f'C{max_line + 2}'].font = Font(bold=True)
        sheet[f'D{max_line + 2}'].font = Font(bold=True)
        sheet[f'A{max_line + 3}'].font = Font(bold=True)
        sheet[f'A{max_line + 4}'].font = Font(bold=True)

        start = constants.args.start[i]
        end = constants.args.end[i]

        avg_emergency_flows_delay, avg_non_emergency_flows_delay = get_flow_delays(start, end)

        # Define the row range of data to consider
        row_range = constants.last_line_data

        # Set the formula for the Non-Emergency Flows
        sheet[f'B{max_line + 3}'] = f'=ROUND(AVERAGEIF(E1:E{row_range}, "<40" , O1:O{row_range}), 2'  
        sheet[f'B{max_line + 4}'] = avg_non_emergency_flows_delay

        # Set the formula for the Emergency Flows
        sheet[f'C{max_line + 3}'] = f'=ROUND(AVERAGEIF(E1:E{row_range}, ">=40", O1:O{row_range}), 2'
        sheet[f'C{max_line + 4}'] = avg_emergency_flows_delay

        #Set comparasion formulas, for the AVG 1º Packet Delay and AVG Flow Delay in percentage
        sheet[f'D{max_line + 3}'] = f'=IFERROR(ROUND((C{max_line + 3} - B{max_line + 3})/ABS(B{max_line + 3}) * 100, 2), "none")'
        sheet[f'D{max_line + 4}'] = f'=IFERROR(ROUND((C{max_line + 4} - B{max_line + 4})/ABS(B{max_line + 4}) * 100, 2), "none")'


    workbook.save(constants.final_file_path)

def set_copied_values(sheet, test_case, start_line):    
    print("Seting values copy from other sheets")
    
    # Cycle through the variables to compare (lines)
    for variable_number in range(constants.num_values_to_compare_all_tests):
        
        # Cycle through the args.f to copy the values (columns)
        for i in range(len(constants.args.f)):
            
            #--------------Collumn C is the second algorithm
            #parse 1st element pre _ in args.f
            sheet_to_copy_from_name = constants.args.f[i].split("_")[0]
            line, column = get_line_column_to_copy_from(sheet_to_copy_from_name, variable_number)

            if line is None or column is None:
                print(f"Error getting line and column to copy from, sheet_to_copy_from: {sheet_to_copy_from_name}, variable number: {variable_number}")
                continue

            cell_reference = f"{column}{line}"
            formula = f"='{sheet_to_copy_from_name}'!{cell_reference}"
            sheet[f'{get_column_letter(2 + i)}{start_line + variable_number + 1}'] = formula

def set_Comparison_sheet():
    print("Setting the Comparison sheet")

    # Create the comparison sheet
    workbook = load_workbook(constants.final_file_path)
    sheet = workbook.create_sheet(title="Comparison")

    title = "Load Test Cases"
    sheet[f'A1'] = title
    sheet[f'A1'].font = Font(bold=True)

    sheet[f'A2'] = f"Variation: From {constants.algorithms[0]} to {constants.algorithms[1]}"

    # Empty line
    sheet.append([""])
    
    # Create a block for each test case
    for test_case in constants.test_cases:
        # Get max line considering the previous test cases
        max_line = sheet.max_row + 1

        set_algorithm_headers(sheet, test_case, max_line)
        set_comparasion_formulas(sheet, max_line)
        set_copied_values(sheet, test_case, max_line)

        # Insert 2 empty lines
        sheet.append([""])
        sheet.append([""])

    # Save the workbook
    workbook.save(constants.final_file_path)


def configure_final_file():
    constants.get_all_sorted_DSCP()

    # Raw data area
    set_pkt_loss()
    set_fist_pkt_delay()

    # Calculations area for each dscp 
    set_caculation_section(-1)             #All Flows
    for dscp in constants.All_DSCP:        #Each DSCP
        set_caculation_section(dscp)

    set_compare_non_Emergency_to_Emergency_variation()
    set_Comparison_sheet()



def set_algorithm_headers(sheet, test_case, start_line):

    # Set test case name in bold test
    title = f"{test_case}"
    sheet[f'A{start_line}'] = title
    sheet[f'A{start_line}'].font = Font(bold=True)

    # Set the collumn names
    sheet[f'B{start_line}'] = constants.algorithms[0]
    sheet[f'C{start_line}'] = constants.algorithms[1]
    sheet[f'D{start_line}'] = "Variation (%)"

    # Set collumn names in bold text
    sheet[f'B{start_line}'].font = Font(bold=True)
    sheet[f'C{start_line}'].font = Font(bold=True)
    sheet[f'D{start_line}'].font = Font(bold=True)

    # Set the lines names
    sheet[f'A{start_line + 1}'] = constants.headers_lines[0]
    sheet[f'A{start_line + 2}'] = constants.headers_lines[1]
    sheet[f'A{start_line + 3}'] = constants.headers_lines[2]
    sheet[f'A{start_line + 4}'] = constants.headers_lines[3]
    sheet[f'A{start_line + 5}'] = constants.headers_lines[4]
    sheet[f'A{start_line + 6}'] = constants.headers_lines[5]
    sheet[f'A{start_line + 7}'] = constants.headers_lines[6]
    sheet[f'A{start_line + 8}'] = constants.headers_lines[7]
    sheet[f'A{start_line + 9}'] = constants.headers_lines[8]
    sheet[f'A{start_line + 10}'] = constants.headers_lines[9]
    sheet[f'A{start_line + 11}'] = constants.headers_lines[10]
    sheet[f'A{start_line + 12}'] = constants.headers_lines[11]
    sheet[f'A{start_line + 13}'] = constants.headers_lines[12]
    sheet[f'A{start_line + 14}'] = constants.headers_lines[13]
    sheet[f'A{start_line + 15}'] = constants.headers_lines[14]
    sheet[f'A{start_line + 16}'] = constants.headers_lines[15]


    # Set lines names in bold text
    sheet[f'A{start_line + 1}'].font = Font(bold=True)
    sheet[f'A{start_line + 2}'].font = Font(bold=True)
    sheet[f'A{start_line + 3}'].font = Font(bold=True)
    sheet[f'A{start_line + 4}'].font = Font(bold=True)
    sheet[f'A{start_line + 5}'].font = Font(bold=True)
    sheet[f'A{start_line + 6}'].font = Font(bold=True)
    sheet[f'A{start_line + 7}'].font = Font(bold=True)
    sheet[f'A{start_line + 8}'].font = Font(bold=True)
    sheet[f'A{start_line + 9}'].font = Font(bold=True)
    sheet[f'A{start_line + 10}'].font = Font(bold=True)
    sheet[f'A{start_line + 11}'].font = Font(bold=True)
    sheet[f'A{start_line + 12}'].font = Font(bold=True)
    sheet[f'A{start_line + 13}'].font = Font(bold=True)
    sheet[f'A{start_line + 14}'].font = Font(bold=True)
    sheet[f'A{start_line + 15}'].font = Font(bold=True)
    sheet[f'A{start_line + 16}'].font = Font(bold=True)

def set_comparasion_formulas(sheet, start_line):
    # Set the formulas to compare the results between the test cases
    for i in range(1, constants.num_values_to_compare_all_tests + 1):
        #print(sheet[f'A{start_line + i}'].value)
        sheet[f'D{start_line + i}'] = f'=IFERROR(ROUND((C{start_line + i} - B{start_line + i}) / ABS(B{start_line + i}) * 100, 2), 0)'
